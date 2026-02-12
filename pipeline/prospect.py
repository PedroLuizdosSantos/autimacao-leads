from pathlib import Path
import unicodedata

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_COLUMNS = [
    "priority_score",
    "fonte",
    "nome",
    "instagram",
    "telefone",
    "whatsapp_link",
    "site",
    "cidade",
    "pais",
    "bairro",
    "nicho",
    "tem_link_na_bio",
    "tem_site",
    "tem_whatsapp_visivel",
    "score",
    "observacao",
    "link_origem",
    "status",
    "data_coleta",
]


def _normalize(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value).strip().lower())
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _is_yes(value: str) -> bool:
    return _normalize(value) == "sim"


def _is_no(value: str) -> bool:
    return _normalize(value) == "nao"


def _priority_score(row: pd.Series) -> int:
    score = 0

    fonte = str(row.get("fonte", "")).strip().lower()
    status = _normalize(row.get("status", ""))
    telefone = str(row.get("telefone", "")).strip()
    instagram = str(row.get("instagram", "")).strip()
    bairro = str(row.get("bairro", "")).strip()

    if fonte == "instagram" and _is_no(row.get("tem_link_na_bio", "")):
        score += 10
    if fonte in {"maps", "google_maps"} and _is_no(row.get("tem_site", "")):
        score += 7
    if _is_yes(row.get("tem_whatsapp_visivel", "")):
        score += 5
    if telefone:
        score += 3
    if instagram:
        score += 2
    if bairro:
        score += 2
    if status == "novo":
        score += 2
    if status == "ignorar":
        score -= 5
    if status == "enviado":
        score -= 3
    if status == "fechou":
        score -= 10

    return score


def _to_int(value) -> int:
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return 0


def _export_xlsx(df: pd.DataFrame, output_xlsx: Path) -> None:
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="prospeccao")
        ws = writer.sheets["prospeccao"]

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for idx, col in enumerate(df.columns, start=1):
            max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()])
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

        priority_col_index = df.columns.get_loc("priority_score") + 1
        fill = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
        for row_idx in range(2, len(df) + 2):
            value = ws.cell(row=row_idx, column=priority_col_index).value
            if _to_int(value) >= 18:
                ws.cell(row=row_idx, column=priority_col_index).fill = fill


def generate_prospect_from_df(
    df: pd.DataFrame,
    output_csv: Path | str = "prospeccao.csv",
    output_xlsx: Path | str = "prospeccao.xlsx",
    top_n: int = 50,
    min_priority: int = 10,
):
    output_csv_path = Path(output_csv)
    output_xlsx_path = Path(output_xlsx)

    if df is None or df.empty:
        df_empty = pd.DataFrame(columns=OUTPUT_COLUMNS)
        df_empty.to_csv(output_csv_path, index=False, encoding="utf-8-sig")
        _export_xlsx(df_empty, output_xlsx_path)
        return {"selected": 0, "total": 0, "csv": str(output_csv_path), "xlsx": str(output_xlsx_path)}

    work = df.copy()
    work["priority_score"] = work.apply(_priority_score, axis=1)
    work["score_num"] = work["score"].apply(_to_int)
    work["status_norm"] = work["status"].apply(_normalize)
    work["nome_norm"] = work["nome"].astype(str).str.lower()

    filtered = work[(work["status_norm"] == "novo") & (work["priority_score"] >= int(min_priority))].copy()
    filtered = filtered.sort_values(
        by=["priority_score", "score_num", "nome_norm"], ascending=[False, False, True]
    ).head(top_n)

    for col in OUTPUT_COLUMNS:
        if col not in filtered.columns:
            filtered[col] = ""

    output_df = filtered[OUTPUT_COLUMNS].copy()
    output_df.to_csv(output_csv_path, index=False, encoding="utf-8-sig")
    _export_xlsx(output_df, output_xlsx_path)

    return {
        "selected": len(output_df),
        "total": len(work),
        "csv": str(output_csv_path),
        "xlsx": str(output_xlsx_path),
    }


def generate_prospect_list(
    input_csv: Path | str,
    output_csv: Path | str = "prospeccao.csv",
    output_xlsx: Path | str = "prospeccao.xlsx",
    top_n: int = 50,
    min_priority: int = 10,
):
    input_path = Path(input_csv)
    output_csv_path = Path(output_csv)
    output_xlsx_path = Path(output_xlsx)

    if not input_path.exists():
        df_empty = pd.DataFrame(columns=OUTPUT_COLUMNS)
        df_empty.to_csv(output_csv_path, index=False, encoding="utf-8-sig")
        _export_xlsx(df_empty, output_xlsx_path)
        return {"selected": 0, "total": 0, "csv": str(output_csv_path), "xlsx": str(output_xlsx_path)}

    df = pd.read_csv(input_path, dtype=str, keep_default_na=False)
    return generate_prospect_from_df(
        df,
        output_csv_path,
        output_xlsx_path,
        top_n=top_n,
        min_priority=min_priority,
    )

